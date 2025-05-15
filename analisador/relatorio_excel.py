from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

class ExcelReport:
    @staticmethod
    def gerar(caminho_saida, analises):
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumo Artigos"

        headers = ["Artigo", "Nanopartículas", "Utilidades", "Patentes", "Tamanho", "Técnica", "Composição", "Aplicações", "Qtd. NPs", "Qtd. Aplicações", "Qtd. Técnicas"]
        ws.append(headers)

        for entrada in analises:
            d = entrada["dados"]
            ws.append([
                entrada["titulo"], d["nanopartículas"], d["utilidades"],
                d["patentes"], d["tamanho"], d["técnica"],
                d["composição"], d["aplicações"],
                d["np_count"], d["app_count"], d["tech_count"]
            ])

        # Gráficos
        chart_titles = ["Qtd. de Nanopartículas", "Qtd. de Aplicações", "Qtd. de Técnicas"]
        col_start = 9

        for i, title in enumerate(chart_titles):
            chart = BarChart()
            chart.title = title
            data = Reference(ws, min_col=col_start + i, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, f"L{2 + i*16}")

        wb.save(caminho_saida)
